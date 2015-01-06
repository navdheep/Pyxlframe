'''
Created on 05-Jan-2015

@author: naveen
'''
'''''1. from openpyxl import load_workbook
wb2 = load_workbook('/home/naveen/Desktop/output/demo.xlsx')
print wb2.get_sheet_names()
['Sheet2', 'New Title', 'Sheet1']

Ans =['Config', 'Sheet1']'''''
from openpyxl.styles import Style, PatternFill, Border, Side, Alignment, Protection, Font
s = Style(font=Font(name='Calibri',
...                 size=11,
...                 bold=False,
...                 italic=False,
...                 vertAlign=None,
...                 underline='none',
...                 strike=False,
...                 color='FF000000'),
...       fill=PatternFill(fill_type=None,
...                 start_color='FFFFFFFF',
...                 end_color='FF000000'),
...       border=Border(left=Side(border_style=None,
...                                color='FF000000'),
...                      right=Side(border_style=None,
...                                 color='FF000000'),
...                      top=Side(border_style=None,
...                               color='FF000000'),
...                      bottom=Side(border_style=None,
...                                  color='FF000000'),
...                      diagonal=Side(border_style=None,
...                                    color='FF000000'),
...                      diagonal_direction=0,
...                      outline=Side(border_style=None,
...                                   color='FF000000'),
...                      vertical=Side(border_style=None,
...                                    color='FF000000'),
...                      horizontal=Side(border_style=None,
...                                     color='FF000000')),
...     alignment=Alignment(horizontal='general',
...                         vertical='bottom',
...                         text_rotation=0,
...                         wrap_text=False,
...                         shrink_to_fit=False,
...                         indent=0),
...     number_format='General',
...     protection=Protection(locked='inherit',
...                           hidden='inherit'))
>>>